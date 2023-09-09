import cv2
from PIL import Image
from scipy import signal,ndimage
import numpy as np
import matplotlib as mpl
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import math
import openpyxl 
import pickle
import pandas as pd
import os as os
import tiffile as tif
from colormath.color_objects import LabColor, sRGBColor
from colormath.color_conversions import convert_color
import numpy.matlib as matlib


import colorfunctions
from colorfunctions import Sample
from colorfunctions import crop_box, RRC, segment_on_dt, water, time_names_0, output_for_labeling, cc_crop
from colorfunctions import CC_RGB, convert_LAB_RGB, color_calibration, Results
# %matplotlib widget
import matplotlib.patches as patches 
import compextractor
import matplotlib.animation as animation
from colorfunctions import point_ave,Edge_detect,warp_image,clean_warp

def unwrap_tiff(): 
    ID_path = './Sample_set_simple/sample_names.xlsx'
    wb_obj = openpyxl.load_workbook(ID_path) 
    sheet_obj = wb_obj.active 
    Sample_ID = sheet_obj.cell(row=2, column = 1).value
    Notes = sheet_obj.cell(row=2,column=4).value
    Time_Step = sheet_obj.cell(row=2,column=5).value
    pixel_to_length = sheet_obj.cell(row=2,column=3).value
    start_image = sheet_obj.cell(row=2,column=2).value
    cut = sheet_obj.cell(row=2,column=6).value
    start_comp = sheet_obj.cell(row=2,column=7).value
    end_comp = sheet_obj.cell(row=2,column=8).value
    num_pics,time,name = time_names_0(Time_Step,cut,start_image)
    img_name = name[0]

    pickle.dump(num_pics, open('./Sample_set_simple/num_pics','wb'))
    pickle.dump(time, open('./Sample_set_simple/time','wb'))
    pickle.dump(name, open('./Sample_set_simple/name','wb'))
    pickle.dump(start_comp, open('./Sample_set_simple/start_comp','wb'))
    pickle.dump(end_comp, open('./Sample_set_simple/end_comp','wb'))
    return sheet_obj, Sample_ID, Notes, Time_Step, pixel_to_length, start_image, cut, start_comp, num_pics,time,name,img_name


def show_cal_image(img_name,verbose=False):
    # Define Image location
    img_path = f'./Images_simp/{img_name}'
    # Open Image
    img_pil = Image.open(img_path, 'r')
    im = Image.fromarray(np.array(img_pil, dtype=np.uint8), 'RGB')
    # Display the chosen image 
    fig, ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
    ax.imshow(im)
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    plt.title('Image Used For Segmentation')
    if verbose:
        plt.show()
    plt.close()
    return img_path, im

def crop_im(im, image_crop_params, img_path,verbose=False):
    # These xvals and yvals are crop indices for the min and max x value and the min and max y vals
    # respectively
    # Change xvals and yvals to fit the red box around the area to be cropped (should include the glass slide
    # only)
    # Display the Crop Box as a red rectangle over the image
    ll_x, ll_y, width, height = crop_box(im, image_crop_params['xvals'],image_crop_params['yvals'],verbose)
    #ll_x,LL_y, etc. are parameters for croping the image that are calculated from xvals and yvals 
    fig, ax = plt.subplots(1, figsize=(1248 / 200, 1024 / 300))
    crop_params = {
        'theta': 0,
        'x1': ll_x,
        'x2': ll_x+width,
        'y1': ll_y, 
        'y2': ll_y+height
    }

    # Crop the Image
    crop_image = RRC(img_path, crop_params,'cv2')
    crop_image = cv2.cvtColor(crop_image, cv2.COLOR_BGR2RGB)
    crop_image_pil = RRC(img_path, crop_params,'pil')

    # Display the cropped image
    ax.imshow(crop_image_pil)
    plt.title('Cropped Image') 
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    if verbose:
        plt.show()
    plt.close()
    return crop_image, crop_image_pil,crop_params

def segment(crop_image,seg_params,verbose=False):
    # Segment the Droplets
    droplet_count = water(crop_image, seg_params['small_elements_pixels'], seg_params['large_elements_pixels'], seg_params['ks'])
    droplet_count = droplet_count.astype('uint8')

    # Display the Segmentation Results 
    fig, ax = plt.subplots(1,2, figsize=(1248 / 200, 1024 / 300))
    ax[0].imshow(droplet_count)
    ax[1].imshow(crop_image)
    ax[0].title.set_text('Droplet Segmentation Results')
    ax[1].title.set_text('Cropped Image')
    ax[0].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[1].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    if verbose:
        plt.show()
    plt.close()
    pickle.dump(droplet_count, open('./Sample_set_simple/droplet_count','wb'))
    return droplet_count 

def clean(clean_params, seg_params, droplet_count,crop_image,crop_image_pil,verbose=False):
    k = clean_params['k']
    kd = clean_params['kd']
    ks = seg_params['ks']
    large_elements_pixels = seg_params['large_elements_pixels'] 
    small_elements_pixels = seg_params['small_elements_pixels']
    
     
    kernel = np.ones((k,k), np.uint8)
    kerneld = np.ones((kd,kd), np.uint8)
    img_erode = cv2.erode(droplet_count, kernel)
    img_erode = cv2.dilate(img_erode,kerneld)
    
    fig, ax = plt.subplots(1,2, figsize=(1248 / 200, 1024 / 300))
    ax[0].imshow(img_erode)
    ax[0].title.set_text('Post-Processed Droplets') 
    ax[1].imshow(droplet_count)
    ax[1].title.set_text('Original Segmented Image') 
    ax[0].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[1].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    if verbose:
        plt.show()
    plt.close()

    water_params = [large_elements_pixels,small_elements_pixels, ks,k,kd] 
    pickle.dump(water_params, open('./Sample_set_simple/water_params','wb'))
 
    crop_image_erode = np.array(crop_image, dtype=np.uint8)
    PIL_crop_image_erode = np.array(crop_image_pil, dtype=np.uint8)
    img_for = img_erode

    #Super impose the images 
    for n in np.unique(img_for):
        if n != 0:
            x = np.where(img_erode == n)
            crop_image_erode[x] = [255]
            PIL_crop_image_erode[x] = [255]

    #Display Superimposed Image
    fig, ax = plt.subplots(1,1, figsize=(1248 / 200, 1024 / 300))
    ax.imshow(PIL_crop_image_erode)
    ax.title.set_text('Droplets vs Original Image') 
    plt.savefig(f'./figs/Superimposed_watershed_fig', dpi=300, bbox_inches='tight')
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    if verbose:
        plt.show()
    plt.close()
    return img_erode,crop_image_erode, PIL_crop_image_erode

def get_xrite(xrite_crop_params,verbose=False):
    # Xrite crop 
    image = Image.open('./Images_simp/xrite.jpg', 'r')
    img_xrite = Image.fromarray(np.array(image, dtype=np.uint8), 'RGB')
    x_vals, y_vals = xrite_crop_params['xvals'],xrite_crop_params['yvals'] 
    wid_x = xrite_crop_params['wid_x']
    hei_x = xrite_crop_params['hei_x']
    # Also change the values in this functions to move the boxes around
    a = xrite_crop_params['a']
    b = xrite_crop_params['b']
    c = xrite_crop_params['c']
    d = xrite_crop_params['d']
    xc_ll_x, xc_ll_y = cc_crop(x_vals,y_vals,wid_x,hei_x,img_xrite,a,b,c,d,verbose)
    other_crop_params_x= [a,b,c,d]
    pickle.dump(x_vals, open('./Sample_set_simple/x_vals_xrite','wb'))
    pickle.dump(y_vals, open('./Sample_set_simple/y_vals_xrite','wb'))
    pickle.dump(xc_ll_x, open('./Sample_set_simple/xc_ll_x','wb'))
    pickle.dump(xc_ll_y, open('./Sample_set_simple/xc_ll_y','wb'))
    pickle.dump(wid_x, open('./Sample_set_simple/wid_x','wb'))
    pickle.dump(hei_x, open('./Sample_set_simple/hei_x','wb'))
    pickle.dump(other_crop_params_x, open('./Sample_set_simple/other_crop_params_x','wb'))
    return xc_ll_x,xc_ll_y

def load_pick():
    # Extract all the Results from Color Calibration for Data Analysis 
    dbfile = open('./Results/RGB_results_f', 'rb')     
    RGB_result = pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/Number_of_drops', 'rb')     
    Number_of_drops = pickle.load(dbfile)

    dbfile = open('./Sample_set_simple/num_pics', 'rb')     
    num_pics= pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/time', 'rb')     
    time= pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/name', 'rb')     
    name= pickle.load(dbfile)

    dbfile = open('./Sample_set_simple/drops', 'rb')     
    drops= pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/droplet_count', 'rb')     
    droplet_count0 = pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/drop_IDs', 'rb')     
    drop_IDs= pickle.load(dbfile)

    dbfile = open('./Sample_set_simple/img_erode', 'rb')     
    img_erode = pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/crop_params', 'rb')     
    crop_params = pickle.load(dbfile)

    dbfile = open('./Sample_set_simple/start_comp', 'rb')     
    start_comp = pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/end_comp', 'rb')     
    end_comp = pickle.load(dbfile)
    dbfile = open('./Sample_set_simple/sample_ID', 'rb')     
    sample_ID = pickle.load(dbfile)

    # Check if these variables exist if so save them and set a boolean to True 
    already_processed = False
    cv_params = []
    extract_params = []
    if os.path.isfile('./Results/Ic_reorder'):
        dbfile = open('./Sample_set_simple/cv_params', 'rb')     
        cv_params = pickle.load(dbfile)
        dbfile = open('./Sample_set_simple/extract_params', 'rb')     
        extract_params = pickle.load(dbfile)
        already_processed = True
    return RGB_result, Number_of_drops, num_pics, time, name, drops, droplet_count0, drop_IDs, img_erode, crop_params, start_comp, end_comp, sample_ID, already_processed, cv_params,extract_params 

def main_edge_detect(name , crop_params, cv_params, already_processed, detect_params, save_new_params=True, verbose=False):
    # Get cropped Image 
    img_name = name[0]
    img_path = f'./Images_simp/{img_name}'
    crop_image = RRC(img_path, crop_params,'cv2')
    crop_image = cv2.cvtColor(crop_image, cv2.COLOR_BGR2RGB)
    fig,ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
    ax.imshow(crop_image)
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    plt.close()
    # This section finds the edges of the glass slide using computer vision 
    # Check if this data set has been analyzed before
    if already_processed:
        k = cv_params[0]
        ap_size = cv_params[1]
        lb = cv_params[2]
        ub = cv_params[3]
        c = cv_params[4]
        block = cv_params[5]
        ts = cv_params[6]
        srn = cv_params[7]
        stn = cv_params[8]
        min_dis = cv_params[9]
    else:
        k = detect_params['k'] # kernel size for erode and dilating the original image to get rid of image noise  
        ap_size = detect_params['ap_size']
        lb = detect_params['lb']
        ub = detect_params['ub']
        c = detect_params['c']
        block = detect_params['block']
        ts = detect_params['ts']
        srn = detect_params['srn'] 
        stn = detect_params['stn'] 
        min_dis = detect_params['min_dis']

    pt_xy = Edge_detect(crop_image,ap_size,k,lb,ub,c,block,ts,srn,stn,verbose)
    
    ## Get rid of points that are too close to each other
    corners = [1,2,3,4,5] # random array of length greater than four
    starting_pts = np.copy(pt_xy)
    if save_new_params:
        cv_params = [k,ap_size,lb,ub,c,block,ts,srn,stn,min_dis]
        pickle.dump(cv_params, open('./Sample_set_simple/cv_params','wb'))
    
    while len(corners)>4:
        corners_averaged,close_points,corners  = point_ave(min_dis,starting_pts,verbose)
        starting_pts = corners
    print('Found Corners')

    #### Get all four corners of the glass slide 
    fig, ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
    ax.imshow(crop_image)
    ax.scatter(corners[:,0],corners[:,1], marker='o')
    ax.title.set_text('Corners of the Glass Slide')
    fig.tight_layout(pad=1.0)
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    if verbose:
        plt.show()
    plt.close()
    return crop_image, corners,cv_params 

def main_warp(corners,crop_image,img_erode,drop_IDs,Number_of_drops,verbose=False):
    warp = warp_image(corners, crop_image, img_erode)
    fig,ax = plt.subplots(2,figsize=(1248 / 200, 1024 / 300))
    ax[1].imshow(warp)
    ax[1].title.set_text('Warped Image')
    ax[0].imshow(img_erode)
    ax[0].title.set_text('Original Eroded Droplet Image')
    ax[0].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[1].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    fig.tight_layout(pad=1.0)
    if verbose:
        plt.show()
    plt.close()
    # Clean the Warped image to get rid of artifacts from the image manipulation
    final_warp,deconst_warp,warp_clean2 = clean_warp(warp,drop_IDs, Number_of_drops,img_erode)
    fig,ax = plt.subplots(2,figsize=(1248 / 200, 1024 / 300))
    ax[0].imshow(warp_clean2)
    ax[0].title.set_text('Random Pixels Deleted, Image Eroded and Dilated')
    ax[1].imshow(final_warp)
    ax[1].title.set_text('After Removing Noise') 
    ax[0].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[1].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    fig.tight_layout(pad=1.0)
    plt.close()
    # print(drop_IDs)
    # print(np.unique(final_warp).astype('uint8'))
    return warp,final_warp,deconst_warp,warp_clean2

def warp_check(drop_index, final_warp, img_erode, drop_IDs, verbose=False):
    i = drop_index['warp_id'] # Droplet Index (not the same as droplet ID, idex = ID location in drop_IDs array) for Warped Image 
    j = drop_index['og_id']  # Droplet Index for Original Image
    droplet_count = np.copy(final_warp)
    droplet_count[np.where(droplet_count != drop_IDs[i])] = 0
    droplet_count[np.where(droplet_count == drop_IDs[i])] = 255 
    
    if i >=1:
        droplet_count[np.where(droplet_count == drop_IDs[i-1])] = 1
    if i >=2:
        droplet_count[np.where(droplet_count == drop_IDs[i-2])] =  drop_IDs[i-2]
    
    droplet_count2 = np.copy(img_erode)
    droplet_count2[np.where(droplet_count2 == drop_IDs[j])] = 255
    
    if j >=1:
        droplet_count2[np.where(droplet_count2 == drop_IDs[j-1])] = 1
    if j >=2:
        droplet_count2[np.where(droplet_count2 == drop_IDs[j-2])] =  drop_IDs[j-2]
    
    droplet_count3 = np.copy(final_warp)
    droplet_count3[np.where(droplet_count3 == drop_IDs[i])] = 255 
    if i >=1:
        droplet_count3[np.where(droplet_count3 == drop_IDs[i-1])] = 1
    if i >=2:
        droplet_count3[np.where(droplet_count3 == drop_IDs[i-2])] =  drop_IDs[i-2]

    fig, ax = plt.subplots(3,figsize=(1248 / 200, 1024 / 300))   
    ax[0].imshow(droplet_count)
    ax[0].title.set_text(f'Isolated Droplet {i} in Warped Image')
    ax[1].imshow(droplet_count3)
    ax[1].title.set_text(f'Droplet {i} in Warped Image')
    ax[2].imshow(droplet_count2)
    ax[2].title.set_text(f'Droplet {j} in Original Image')
    ax[0].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[1].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[2].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    fig.tight_layout(pad=1.0)
    if verbose:
        plt.show()
    plt.close()

def main_comp(comp_params,already_processed,final_warp,extract_params, save_new_params=True,verbose=False):
    # Read motor speed and gcode files
    with open(os.getcwd()+'/data/motor_speeds.txt', 'r') as file:
        motor_speeds = file.read()
    gcode = pd.read_csv(os.getcwd()+'/data/gcode_XY.csv')

    # Define parameters 
    if already_processed:
        rasterSpeed = extract_params[0] # mm/s gocde raster speed
        rasterScaleY = extract_params[1] # scale gcode raster pattern in Y-dim
        rasterScaleX = extract_params[2] # scale gcode raster pattern in X-dim
        rasterOffsetY = extract_params[3]# offset gcode raster pattern in X-dim
        rasterOffsetX = extract_params[4]# offset gcode raster pattern in X-dim
    else:
        rasterSpeed = comp_params['rasterSpeed'] # mm/s gocde raster speed
        rasterScaleY = comp_params['rasterScaleY'] # scale gcode raster pattern in Y-dim
        rasterScaleX = comp_params['rasterScaleX'] # scale gcode raster pattern in X-dim
        rasterOffsetY = comp_params['rasterOffsetY'] # offset gcode raster pattern in X-dim
        rasterOffsetX = comp_params['rasterOffsetX'] # offset gcode raster pattern in X-dim

    # Need to inverse the image before putting it into the extractor to stop it from being flipped 
    sgmnt = final_warp[::-1]

    # Extract composition of segmented data using motor speeds and gcode
    idx, comp = compextractor.get_compositions(segmentation=sgmnt, motor_speeds=motor_speeds, gcode=gcode, rasterSpeed=rasterSpeed, rasterScaleY=rasterScaleY, rasterScaleX=rasterScaleX, rasterOffsetX=rasterOffsetX, rasterOffsetY=rasterOffsetY, savepath=f'{os.getcwd()}/figs/',verbose=verbose)
    if save_new_params:
        extract_params = [rasterSpeed,rasterScaleY, rasterScaleX,rasterOffsetY,rasterOffsetX]
        pickle.dump(extract_params,open('./Sample_set_simple/extract_params','wb')) 
    # Order is an array where the droplet IDs are organized from the start to the end of the gradient.
    order = np.copy(idx)
    pickle.dump(order, open('./Sample_set_simple/order','wb') )
    compositions = np.copy(comp)
    np.savetxt("./Results/compositions.csv", compositions, delimiter=",")
    pickle.dump(compositions, open('./Sample_set_simple/compositions','wb') )
    return extract_params, order, compositions

def order_check(Number_of_drops,img_erode,order,verbose=False):
    fig, ax = plt.subplots()

    ims = []
    # Increment i (0 to max number of droplets -1) to move from the start to the end of the gradient
    for i in range(Number_of_drops):
        droplet_count = np.copy(img_erode)
        droplet_count[np.where(droplet_count == order[i])] = 255
        # ax.title.set_text(f'Droplet of composition:{np.around(1-comp[i],2)}A to {np.around(comp[i],2)}B. Number:{i+1} of {np.size(order)}')
        im = ax.imshow(droplet_count)
        plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
        ims.append([im])

    ani = animation.ArtistAnimation(fig, ims, interval=100, blit=True, repeat=False)
    #ani = animation.ArtistAnimation(fig, ims, interval=50, blit=True, repeat_delay=1000)
    ani.save('ordered_drops.gif')
    if verbose:
        plt.show()

    
def extract_ic(RGB_result, Number_of_drops,drops,num_pics,verbose=False):
    Test_array = np.array([RGB_result.astype('uint8')]* Number_of_drops)
    index = np.where(drops[:] == 0)
    Test_array[index[0],:,:,index[1],index[2]] = 0
    Test_array_RGB = np.array([[None,None,None]]*num_pics) 
    Test_array_RGB = np.array([Test_array_RGB] * Number_of_drops)
    index = np.where(drops[:] != 0)
    # sum the number of pixels in each droplet 
    drop_pix = []
    for i in range(Number_of_drops):
        drop_pix.append(np.size(np.where(drops[i]!=0),1))
    test = np.sum(Test_array[:,:,:,index[1],index[2],0], axis = 3) # 75 69 1
    r = np.size(test.astype(int),0)
    c = np.size(test.astype(int),1)
    Test_layer = np.zeros((1,r))
    for i in range(c):
        # print(i)
        Test_row = np.divide(test[:,i,:].reshape(r),drop_pix)
        # print(Test_layer)
        if i == 0:
            Test_layer = np.append(Test_layer.reshape(1,r), Test_row.reshape(1,r), axis = 0)
            Test_layer = Test_layer[1,:]
        else:
            Test_layer = np.append(Test_layer.reshape(i,r), Test_row.reshape(1,r), axis = 0)
    Test_array_RGB[:,:,0] = np.array(Test_layer).astype(int).transpose()

    test = np.sum(Test_array[:,:,:,index[1],index[2],1], axis = 3) # 75 69 1
    Test_layer = np.zeros((1,r))
    for i in range(c):
        # print(i)
        Test_row = np.divide(test[:,i,:].reshape(r),drop_pix)
        # print(Test_layer)
        if i == 0:
            Test_layer = np.append(Test_layer.reshape(1,r), Test_row.reshape(1,r), axis = 0)
            Test_layer = Test_layer[1,:]
        else:
            Test_layer = np.append(Test_layer.reshape(i,r), Test_row.reshape(1,r), axis = 0)
    Test_array_RGB[:,:,1] = np.array(Test_layer).astype(int).transpose()

    test = np.sum(Test_array[:,:,:,index[1],index[2],2], axis = 3) # 75 69 1
    Test_layer = np.zeros((1,r))
    for i in range(c):
        # print(i)
        Test_row = np.divide(test[:,i,:].reshape(r),drop_pix)
        # print(Test_layer)
        if i == 0:
            Test_layer = np.append(Test_layer.reshape(1,r), Test_row.reshape(1,r), axis = 0)
            Test_layer = Test_layer[1,:]
        else:
            Test_layer = np.append(Test_layer.reshape(i,r), Test_row.reshape(1,r), axis = 0)
    Test_array_RGB[:,:,2] = np.array(Test_layer).astype(int).transpose()
    fig, ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
    ax.imshow(np.floor(Test_array_RGB).astype(int))
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    plt.title('Colors of Each Sample Over time (Not Ordered by Composition')
    if verbose:
        plt.show()
    plt.close()
    # For each droplet 
    ## at each time stamp I need to calculate the change in color across a color channel since time step 0 
    # then multiply this diference by the time_step length ( ie 5 minutes) 
    # add this to the previous value 
    # Do this for all three color chanells and add the three intergrals 
    # Test_array_RGB[:,:,0] # all the red values of droplets over each time step 
    # Create an array with all first color values of each droplet 
    # Red 
    C0_R = Test_array_RGB[:,3,0] # lenght 75 
    # Green 
    C0_G= Test_array_RGB[:,3,1] # lenght 75
    # Blue 
    C0_B = Test_array_RGB[:,3,2] # lenght 75


    # Subtract this from each column of Test_array
    Test_C0_Ct_R = Test_array_RGB[:,np.arange(1,c-1),0] - C0_R[..., np.newaxis]
    Test_C0_Ct_G = Test_array_RGB[:,np.arange(1,c-1),1] - C0_G[..., np.newaxis]
    Test_C0_Ct_B = Test_array_RGB[:,np.arange(1,c-1),2] - C0_B[..., np.newaxis]

    # Define the time step 
    time_step = 5 # minutes 

    # Multiply by the time step 
    Test_C0_R_Area = np.absolute(Test_C0_Ct_R * time_step)
    Test_C0_G_Area = np.absolute(Test_C0_Ct_G * time_step)
    Test_C0_B_Area = np.absolute(Test_C0_Ct_B * time_step)

    # Sum along the rows  = Suming all the (C0 -Ct)* time_step Values for all time steps for each droplet 
    Test_sum_int_R = np.sum(Test_C0_R_Area, axis = 1 )
    Test_sum_int_G = np.sum(Test_C0_G_Area, axis = 1 )
    Test_sum_int_B = np.sum(Test_C0_B_Area, axis = 1 )
    # The result of the last step should be 3, 75 long arrays with each element being the intergral sum of the R,G, or B values of that droplet over time 

    Ic = Test_sum_int_R+Test_sum_int_G+Test_sum_int_B
    # The resulting Ic should be of length 75 with the Instability Index for each droplet 
    # Visualize Ic 

    fig , ax = plt.subplots(1)
    plt.plot(Ic,'bo')
    plt.xlabel("Droplet Index")
    plt.ylabel("Instability Index")
    plt.title('Instability Indices of all samples (Not Ordered by Composition)')
    if verbose:
        plt.show()
    plt.close()
    return Ic,Test_array_RGB,c

def order_ic(Ic, Test_array_RGB,Number_of_drops,drop_IDs,order,comp,verbose=False):
    # Reorder Test_array_RGB
    Test_array_reorder_i = np.copy(Test_array_RGB)
    Ic_reorder = np.copy(Ic)
    Test_array_reorder = np.zeros_like(Test_array_reorder_i)
    for i in range(Number_of_drops):
        Test_array_reorder[i,:] = Test_array_reorder_i[np.where(drop_IDs == order[i])[0][0],:].astype(int)
        Ic_reorder[i] = Ic[np.where(drop_IDs == order[i])[0][0]]

    fig, ax = plt.subplots(1,2,figsize=(1248 / 200, 1024 / 300))
    ax[0].imshow(Test_array_reorder.astype(int))
    ax[1].imshow(Test_array_RGB.astype(int))
    ax[0].title.set_text("Reordered")
    ax[1].title.set_text("Original")
    ax[0].set_xlabel("Time Step in 5 Minute Intervals")
    ax[1].set_xlabel("Time Step in 5 Minute Intervals")
    ax[0].set_ylabel("Droplet Index")
    ax[0].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax[1].tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    
    
    plt.savefig(f'./Results/Reordered_fig', dpi=300, bbox_inches='tight')
    # plt.ylabel("Droplet Index")
    if verbose:
        plt.show()
    plt.close()

    fig , ax = plt.subplots(1)
    x = 100*comp
    points = ax.scatter(x, Ic_reorder, c=100*comp, cmap='viridis')
    plt.ticklabel_format(axis='both', style='sci', scilimits=(1,4))
    fig.colorbar(points,label = "%MA")
    ax.grid()
    ax.set_xlabel('% MA Composition',labelpad=20)
    ax.set_ylabel('Instability Index',labelpad=20)
    ax.set_title('Instability of Droplets Along Gradient',pad=20)
    plt.savefig(f'./figs/Instability_Index_fig', dpi=300, bbox_inches='tight')
    if verbose:
        plt.show()
    plt.close()

    pickle.dump(Ic_reorder, open('./Results/Ic_reorder','wb'))
    return Ic_reorder,Test_array_reorder



def plot_grand(num_pics,Test_array_RGB, Test_array_reorder, time, start_comp, end_comp, comp, Number_of_drops, spacing, verbose=False):
    # Now lets Color Correct the data again 
    # Set a median color value 
    R_start = np.mean(Test_array_RGB[:,1,0])
    G_start = np.mean(Test_array_RGB[:,1,1])
    B_start = np.mean(Test_array_RGB[:,1,2])
    # Find the diference between this median and everydroplet at time t =1 
    Correct_R = Test_array_reorder[:,1,0] - R_start
    Correct_G = Test_array_reorder[:,1,1] - G_start
    Correct_B = Test_array_reorder[:,1,2] - B_start

    Test_array_RGB_cor = np.copy(Test_array_reorder)
    Test_array_RGB_cor[:,:,0] = Test_array_RGB_cor[:,:,0] - Correct_R[..., np.newaxis]
    Test_array_RGB_cor[:,:,1] = Test_array_RGB_cor[:,:,1] - Correct_G[..., np.newaxis]
    Test_array_RGB_cor[:,:,2] = Test_array_RGB_cor[:,:,2] - Correct_B[..., np.newaxis]
    pickle.dump(Test_array_RGB_cor, open('./Sample_set_simple/Final_Test_array_f','wb'))

    # Plot the correct results 
    y_int = spacing['y_int']
    x_int = spacing['x_int']
    fig, ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
    ax.imshow(Test_array_RGB_cor.astype(int)) 
    plt.xticks(ticks = np.arange(0,num_pics,num_pics/len(time[::x_int])), labels = np.array(time[::x_int]/60).astype(int))
    Al = np.char.add(f'{start_comp}','$_{')
    Al = np.char.add(Al,np.around(1-comp[::y_int],2).astype(str))
    Al = np.char.add(Al,'}')
    Bl =  np.char.add(f'${end_comp}','$_{')
    Bl =  np.char.add(Bl,np.around(comp[::y_int],2).astype(str))
    Bl = np.char.add(Bl,'}$')
    labely = np.char.add(Al,Bl)
    plt.yticks(ticks = np.arange(0,Number_of_drops,Number_of_drops/len(comp[::y_int])), labels = labely)

    plt.xlabel("Time (min)")
    plt.ylabel("Composition")
    # plt.title("Droplet Colors Over Time")
    plt.savefig('./Results/Grand_Plot', dpi=300, bbox_inches='tight')
    plt.savefig('./figs/Grand_Plot_corrected_labeled', dpi=300, bbox_inches='tight')
    if verbose:
        plt.show()
    plt.close()
    fig,ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
    plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
    ax.imshow(Test_array_reorder.astype(int))
    plt.axis('off')
    plt.savefig('./figs/Grand_Plot_nocorrection', dpi=300, bbox_inches='tight')
    plt.close()
    
def degrade_gif(num_pics,RGB_result,verbose=False):
    fig, ax = plt.subplots()
    ims = []
    for i in range(num_pics):
        im = ax.imshow(RGB_result[i][0].astype(int))
        plt.tick_params(left = False, right = False , labelleft = False ,
                labelbottom = False, bottom = False)
        ims.append([im])

    ani = animation.ArtistAnimation(fig, ims, interval=50, blit=True, repeat=False)
    # ani = animation.ArtistAnimation(fig, ims, interval=50, blit=True, repeat_delay=1000)
    ani.save('degradation.gif')


def individual_drops(c,drop_IDs,Ic_reorder,Test_array_reorder,verbose=False):
    # save Ic-reorder as a csv 
    # a = numpy.asarray([ [1,2,3], [4,5,6], [7,8,9] ])
    np.savetxt("./Results/Ic.csv", Ic_reorder, delimiter=",")
    # Lets save all the 1D plots of the droplets Aleks Wants 
    desi = drop_IDs['desi']
    for drop_num in desi:
        line = Test_array_reorder[drop_num-1,:].astype(int)
        line_print = np.stack((line,line,line,line), axis = 0)
        line_print = line_print[:,np.arange(1,c),:]
        fig, ax = plt.subplots(1,figsize=(1248 / 200, 1024 / 300))
        ax.imshow(line_print)
        plt.title(f'Droplet {drop_num}')
        plt.axis('off')
        plt.savefig(f'./Results/1D_Drop_{drop_num}', dpi=300, bbox_inches='tight')
        plt.savefig(f'./figs/1D_Drop_{drop_num}_fig', dpi=300, bbox_inches='tight')
        if verbose:
            plt.show()
        plt.close()
